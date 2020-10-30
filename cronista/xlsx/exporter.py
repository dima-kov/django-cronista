import logging
from urllib.parse import quote

from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from cronista.base import BaseModelExporter
from cronista.base.base import BaseExporter

logger = logging.getLogger(__name__)


class BaseXlsxExporter(BaseExporter):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.wb = Workbook()
        self.ws = self.wb.active

    def export(self):
        self.header()
        self.export_body()

    def as_http_response(self, filename='export'):
        filename = quote('{}.xlsx'.format(filename))
        response = HttpResponse(
            content=save_virtual_workbook(self.wb),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename={}'.format(filename)
        return response

    def as_file(self, filename='export'):
        self.wb.save(filename)

    def header(self):
        pass

    def export_body(self):
        pass


class XlsxModelExporter(BaseXlsxExporter, BaseModelExporter):
    header_row = 1
    header_col = 1
    default_value = _('Інформація відсутня')

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.define_start_cols()

    def field_exporter_merge_header(self, field_exporter, row):
        self.ws.merge_cells(
            start_row=row, start_column=field_exporter.start_col,
            end_row=row, end_column=field_exporter.get_end_col()
        )

    def field_exporter_header(self, field_exporter, row, col):
        for verbose_name in field_exporter.get_fields_verbose_names():
            self.put_value(verbose_name, row, col)
            col += 1
        return col

    def header(self):
        row = self.header_row
        col = self.header_col
        for field_exporter in self.field_exporters:
            if not field_exporter.multiple:
                # header for fk, single object
                col = self.field_exporter_header(field_exporter, row, col)
                continue

            # header for multiple: merge header_row, place related_field name
            # and place related model fields in next row
            self.field_exporter_merge_header(field_exporter, row)
            self.put_value(self.get_field_verbose_name(field_exporter.name), row, col)
            for _ in range(field_exporter.max_num):
                col = self.field_exporter_header(field_exporter, row + 1, col)

    def field_exporter_data(self, obj, field_exporter, row, col):
        for data in field_exporter.get_data_set(obj):
            self.put_value(data, row, col)
            col += 1
        return col

    def get_data_start_row(self):
        """Returns row from witch data goes"""
        return self.header_row + 2

    def export_body(self):
        data_start_row = self.get_data_start_row()
        self.export_qs(self.qs, data_start_row)

    def export_qs(self, qs, start_row):
        """Performs export of queryset"""
        row = start_row
        for obj in qs:
            self.export_obj(obj, row)
            row += 1
        return row

    def export_obj(self, obj, row):
        """Performs export of one object"""
        col = 1
        for field_exporter in self.field_exporters:
            if field_exporter.multiple:
                # Fields with multiple
                field = getattr(obj, field_exporter.name)
                data = getattr(field, 'all')()

                col = field_exporter.start_col

            elif field_exporter.name:
                # FK, O2O fields
                data = getattr(obj, field_exporter.name)
                col = field_exporter.start_col
            else:
                # local fields exporter
                data = obj

            col = self.field_exporter_data(data, field_exporter, row, col)

    def put_value(self, value, row, col):
        value = value or self.default_value
        self.ws.cell(row=row, column=col, value=str(value))


class XlsxChunkModelExporter(XlsxModelExporter):
    pagination_chunk = 1000

    def export_body(self):
        row = self.get_data_start_row()
        qs = self.get_queryset()
        all_count = qs.count()

        for chunk in range(0, all_count, self.pagination_chunk):
            chunk_qs = qs[chunk:chunk + self.pagination_chunk]
            row = self.export_qs(chunk_qs, row)
            logger.info(f'Progress: {chunk} {all_count}')
