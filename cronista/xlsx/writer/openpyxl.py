from urllib.parse import quote

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.writer.excel import save_virtual_workbook

from cronista.base import ExporterWriter


class OpenPyXlWriter(ExporterWriter):
    default_value = ''

    def __init__(self):
        super().__init__()
        self.wb = Workbook()
        self.ws = self.wb.active

    def write(self, x, y, value):
        value = value or self.default_value
        cell = self.ws.cell(row=y, column=x)
        cell.value = str(value)

    def move_left(self, x_from, steps):
        max_col = self.ws.max_column
        if x_from > self.ws.max_column:
            max_col = x_from

        c = CellRange(
            min_col=x_from,
            min_row=3,
            max_col=max_col,
            max_row=self.ws.max_row
        )
        self.ws.move_range(c, rows=0, cols=steps)

    def duplicate_range(self, min_col, min_row, max_col, max_row, row_shift=0, col_shift=0):
        if max_col is None:
            max_col = self.ws.max_column

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                value = self.ws.cell(row=row, column=col).value
                if not value:
                    continue

                self.write(
                    x=col + col_shift,
                    y=row + row_shift,
                    value=value,
                )

    def merge_range(self, min_col, min_row, max_col, max_row):
        self.ws.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )

    def freeze_panes(self, col, row):
        cell = self.ws.cell(row=row, column=col)
        self.ws.freeze_panes = cell

    def to_file(self, filename='export'):
        self.wb.save(filename)

    def to_binary(self):
        return save_virtual_workbook(self.wb)

    def to_response(self, filename='export'):
        filename = quote('{}.xlsx'.format(filename))
        response = HttpResponse(
            content=save_virtual_workbook(self.wb),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename={}'.format(filename)
        return response
