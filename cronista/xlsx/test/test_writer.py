from unittest import TestCase

from openpyxl import Workbook

from cronista.xlsx.writer import XlsxWriter


class WriterTestCase(TestCase):

    @classmethod
    def setUpClass(cls) -> None:
        cls.writer: XlsxWriter = XlsxWriter()

    def test_writer_max_col_row(self):
        self.assertEqual(self.writer.max_col, 0)
        self.assertEqual(self.writer.max_row, 0)

        self.writer.max_col = 10
        self.assertEqual(self.writer.max_col, 10)
        self.writer.max_col = 5
        self.assertEqual(self.writer.max_col, 10)

        self.writer.max_row = 40
        self.assertEqual(self.writer.max_row, 40)
        self.writer.max_row = 15
        self.assertEqual(self.writer.max_row, 40)
