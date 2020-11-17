from urllib.parse import quote

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.writer.excel import save_virtual_workbook

from cronista.base import BaseExporter
from cronista.base import ModelExporter
from cronista.base.writer import ExportWriter


class XlsxWriter(ExportWriter):
    default_value = '-'

    def __init__(self, ws):
        self.ws = ws
        self._max_col = 0
        self._max_row = 0

    def write(self, x, y, value):
        value = value or self.default_value
        self.ws.cell(row=y, column=x, value=str(value))
        self.max_col = x
        self.max_row = y

    def move_left(self, x_from, steps):
        print('move!!!', x_from, steps)
        c = CellRange(min_col=x_from, min_row=3, max_col=self.max_col, max_row=self.max_row)
        self.ws.move_range(c, rows=0, cols=steps)

    @property
    def max_col(self):
        return self._max_col

    @max_col.setter
    def max_col(self, value):
        if value > self._max_col:
            self._max_col = value

    @property
    def max_row(self):
        return self._max_row

    @max_row.setter
    def max_row(self, value):
        if value > self._max_row:
            self._max_row = value


class BaseXlsxExporter(BaseExporter):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.wb = Workbook()
        self.ws = self.wb.active

    def as_http_response(self, filename='export'):
        filename = quote('{}.xlsx'.format(filename))
        response = HttpResponse(
            content=save_virtual_workbook(self.wb),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename={}'.format(filename)
        return response

    def as_file(self, filename='export'):
        self.wb.save(filename)

    def header(self):
        pass

    def export_body(self):
        pass


class XlsxModelExporter(ModelExporter, BaseXlsxExporter):

    def __init__(self):
        super().__init__()
        self.writer = XlsxWriter(self.ws)

    def export(self, qs):
        super().export(qs, self.writer)
